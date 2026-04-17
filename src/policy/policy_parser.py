from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

import pandas as pd


WORD_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
SUPPORTED_SUFFIXES = (".docx", ".xlsx", ".xls", ".pdf", ".doc")


@dataclass
class PolicyParseResult:
    inventory: pd.DataFrame
    rule_table: pd.DataFrame
    failures: pd.DataFrame


def _list_policy_files(policy_dir: str | Path) -> list[Path]:
    root = Path(policy_dir)
    files: list[Path] = []
    for suffix in SUPPORTED_SUFFIXES:
        files.extend(sorted(path for path in root.rglob(f"*{suffix}") if not path.name.startswith(".")))
    return sorted(set(files))


def _serialize_source_file(path: Path, project_root: Path | None) -> str:
    resolved = path.resolve()
    if project_root is None:
        return resolved.as_posix()
    try:
        return resolved.relative_to(project_root).as_posix()
    except ValueError:
        return resolved.as_posix()


def _extract_docx_text(path: Path) -> str:
    with ZipFile(path) as archive:
        xml_bytes = archive.read("word/document.xml")
    root = ET.fromstring(xml_bytes)
    texts = [node.text.strip() for node in root.findall(".//w:t", WORD_NS) if node.text and node.text.strip()]
    return "".join(texts)


def _extract_xlsx_summary(path: Path) -> dict[str, Any]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True)
    worksheet = workbook.worksheets[0]
    rows = list(worksheet.iter_rows(min_row=1, max_row=300, values_only=True))
    frame = pd.DataFrame(rows[2:], columns=rows[1]).dropna(how="all")
    summary = {
        "sheet_names": "|".join(workbook.sheetnames),
        "mechanism_price_floor": None,
        "mechanism_price_ceiling": None,
        "mechanism_volume_ratio_max": None,
        "mechanism_exec_term_years": None,
        "table_preview": frame.head(10).to_dict(orient="records"),
    }
    if "机制电价（元/千瓦时)" in frame.columns:
        prices = pd.to_numeric(frame["机制电价（元/千瓦时)"], errors="coerce").dropna()
        if not prices.empty:
            summary["mechanism_price_floor"] = float(prices.min())
            summary["mechanism_price_ceiling"] = float(prices.max())
    if "机制电量比例(%)" in frame.columns:
        ratios = pd.to_numeric(frame["机制电量比例(%)"], errors="coerce").dropna()
        if not ratios.empty:
            summary["mechanism_volume_ratio_max"] = float(ratios.max()) / 100.0
    if "机制电价执行期限(年)" in frame.columns:
        years = pd.to_numeric(frame["机制电价执行期限(年)"], errors="coerce").dropna()
        if not years.empty:
            summary["mechanism_exec_term_years"] = float(years.max())
    return summary


def _parse_filename_dates(name: str) -> tuple[pd.Timestamp | None, pd.Timestamp | None]:
    matches = re.findall(r"(20\d{2})[-.年](\d{1,2})[-.月](\d{1,2})", name)
    if not matches:
        return None, None
    timestamps = [pd.Timestamp(year=int(y), month=int(m), day=int(d)) for y, m, d in matches]
    if len(timestamps) == 1:
        return timestamps[0], timestamps[0]
    return timestamps[0], timestamps[1]


def _extract_notice_date(text: str) -> pd.Timestamp | None:
    matches = re.findall(r"(20\d{2})年(\d{1,2})月(\d{1,2})日", text)
    if not matches:
        return None
    year, month, day = map(int, matches[-1])
    return pd.Timestamp(year=year, month=month, day=day)


def _append_rule(
    rows: list[dict[str, Any]],
    *,
    rule_id: str,
    policy_name: str,
    publish_time: pd.Timestamp | None,
    effective_start: pd.Timestamp | None,
    effective_end: pd.Timestamp | None,
    rule_type: str,
    scope: str,
    state_group: str,
    state_name: str,
    state_value: Any,
    source_file: str,
    note: str,
) -> None:
    rows.append(
        {
            "rule_id": rule_id,
            "policy_name": policy_name,
            "publish_time": publish_time,
            "effective_start": effective_start,
            "effective_end": effective_end,
            "rule_type": rule_type,
            "scope": scope,
            "state_group": state_group,
            "state_name": state_name,
            "state_value": state_value,
            "source_file": source_file,
            "note": note,
        }
    )


def _add_rule(
    rows: list[dict[str, Any]],
    *,
    policy_name: str,
    publish_time: pd.Timestamp | None,
    effective_start: pd.Timestamp | None,
    effective_end: pd.Timestamp | None,
    rule_type: str,
    scope: str,
    state_group: str,
    state_name: str,
    state_value: Any,
    source_file: str,
    note: str,
) -> None:
    _append_rule(
        rows,
        rule_id=f"rule_{len(rows)+1:03d}",
        policy_name=policy_name,
        publish_time=publish_time,
        effective_start=effective_start,
        effective_end=effective_end,
        rule_type=rule_type,
        scope=scope,
        state_group=state_group,
        state_name=state_name,
        state_value=state_value,
        source_file=source_file,
        note=note,
    )


def _parse_docx_or_note(path: Path) -> tuple[str, str, str]:
    suffix = path.suffix.lower()
    if suffix == ".docx":
        return _extract_docx_text(path), "ok", ""
    if suffix in {".pdf", ".doc"}:
        return "", "skipped", f"暂不解析 {suffix}"
    return "", "ok", ""


def parse_policy_environment(policy_dir: str | Path, project_root: str | Path | None = None) -> PolicyParseResult:
    policy_root = Path(policy_dir).resolve()
    project_root_path = Path(project_root).resolve() if project_root is not None else policy_root.parent
    files = _list_policy_files(policy_root)
    inventory_rows: list[dict[str, Any]] = []
    rule_rows: list[dict[str, Any]] = []
    failure_rows: list[dict[str, Any]] = []
    mechanism_xlsx_summary: dict[str, Any] = {}
    mechanism_result_file: str | None = None

    for path in files:
        publish_time, effective_start_guess = _parse_filename_dates(path.name)
        effective_start = effective_start_guess
        extracted_text = ""
        parse_status = "ok"
        parse_note = ""

        try:
            if path.suffix.lower() in {".docx", ".pdf", ".doc"}:
                extracted_text, parse_status, parse_note = _parse_docx_or_note(path)
                publish_time = publish_time or _extract_notice_date(extracted_text)
            elif path.suffix.lower() in {".xlsx", ".xls"}:
                mechanism_xlsx_summary = _extract_xlsx_summary(path)
                mechanism_result_file = _serialize_source_file(path, project_root_path)
                parse_note = f"sheet_names={mechanism_xlsx_summary['sheet_names']}"
        except Exception as exc:  # pragma: no cover - 容错记录
            parse_status = "failed"
            parse_note = str(exc)
            failure_rows.append({"source_file": _serialize_source_file(path, project_root_path), "error": str(exc)})

        inventory_rows.append(
            {
                "source_file": _serialize_source_file(path, project_root_path),
                "file_name": path.name,
                "suffix": path.suffix.lower(),
                "publish_time": publish_time,
                "effective_start_guess": effective_start,
                "parse_status": parse_status,
                "parse_note": parse_note,
            }
        )

        if parse_status in {"failed", "skipped"}:
            continue

        policy_name = path.stem
        source_file = _serialize_source_file(path, project_root_path)
        scope = "湖南省售电周度持仓实验"

        if "现货市场交易实施细则" in path.name:
            base_effective = effective_start or publish_time
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=base_effective,
                effective_end=None,
                rule_type="market_coupling",
                scope=scope,
                state_group="market_coupling",
                state_name="lt_settlement_base",
                state_value=1.0,
                source_file=source_file,
                note="中长期动作用作结算底仓，不直接映射物理执行曲线。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=base_effective,
                effective_end=None,
                rule_type="market_coupling",
                scope=scope,
                state_group="market_coupling",
                state_name="spot_marginal_exposure",
                state_value=1.0,
                source_file=source_file,
                note="现货市场承担边际敞口。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=base_effective,
                effective_end=None,
                rule_type="market_coupling",
                scope=scope,
                state_group="market_coupling",
                state_name="lt_spot_coupling_state",
                state_value=1.0,
                source_file=source_file,
                note="中长期仅作结算依据，现货承担全电量集中竞价下的边际风险。",
            )

        if any(keyword in path.name for keyword in ["辅助服务市场建设", "调频辅助服务市场交易实施细则", "辅助服务市场交易规则"]):
            base_effective = effective_start or publish_time
            text = extracted_text or ""
            peak_shaving_paused = float("暂停调峰" in text or "调峰辅助服务暂停" in text)
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=base_effective,
                effective_end=None,
                rule_type="ancillary_coupling",
                scope=scope,
                state_group="ancillary_coupling",
                state_name="ancillary_peak_shaving_pause",
                state_value=peak_shaving_paused,
                source_file=source_file,
                note="辅助服务状态会影响现货边际可调空间。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=base_effective,
                effective_end=None,
                rule_type="ancillary_coupling",
                scope=scope,
                state_group="ancillary_coupling",
                state_name="ancillary_freq_reserve_tight",
                state_value=1.0,
                source_file=source_file,
                note="调频容量预留压缩现货边际可调空间。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=base_effective,
                effective_end=None,
                rule_type="ancillary_coupling",
                scope=scope,
                state_group="ancillary_coupling",
                state_name="ancillary_price_boundary_tight",
                state_value=1.0,
                source_file=source_file,
                note="辅助服务耦合会改变价格波动边界与现货风险窗口。",
            )

        if any(keyword in path.name for keyword in ["信息披露基本规则", "计量实施规则", "消纳监测统计实施细则"]):
            base_effective = effective_start or publish_time
            if "信息披露基本规则" in path.name:
                _add_rule(
                    rule_rows,
                    policy_name=policy_name,
                    publish_time=publish_time,
                    effective_start=base_effective,
                    effective_end=None,
                    rule_type="info_forecast_boundary",
                    scope=scope,
                    state_group="info_forecast_boundary",
                    state_name="info_disclosure_active",
                    state_value=1.0,
                    source_file=source_file,
                    note="信息披露制度改善市场可观测性。",
                )
            if "计量实施规则" in path.name:
                _add_rule(
                    rule_rows,
                    policy_name=policy_name,
                    publish_time=publish_time,
                    effective_start=base_effective,
                    effective_end=None,
                    rule_type="info_forecast_boundary",
                    scope=scope,
                    state_group="info_forecast_boundary",
                    state_name="metering_boundary_improved",
                    state_value=1.0,
                    source_file=source_file,
                    note="计量规则改善偏差结算与观测边界。",
                )
            if "消纳监测统计实施细则" in path.name:
                _add_rule(
                    rule_rows,
                    policy_name=policy_name,
                    publish_time=publish_time,
                    effective_start=base_effective,
                    effective_end=None,
                    rule_type="info_forecast_boundary",
                    scope=scope,
                    state_group="info_forecast_boundary",
                    state_name="forecast_boundary_improved",
                    state_value=1.0,
                    source_file=source_file,
                    note="新能源消纳监测统计提升新能源波动的观测边界。",
                )

        if any(
            keyword in path.name
            for keyword in ["深化新能源上网电价市场化改革", "新能源机制电价竞价实施细则", "新能源机制电价竞价工作有关事项"]
        ):
            stage_effective = effective_start or publish_time
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=stage_effective,
                effective_end=pd.Timestamp("2025-12-31 23:59:59"),
                rule_type="renewable_mechanism_stage",
                scope=scope,
                state_group="renewable_mechanism",
                state_name="mechanism_stage_label",
                state_value="竞价准备",
                source_file=source_file,
                note="新能源机制电价制度进入竞价准备阶段。",
            )
            execution_effective = pd.Timestamp("2026-01-01 00:00:00")
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=execution_effective,
                effective_end=None,
                rule_type="renewable_mechanism_execution",
                scope=scope,
                state_group="renewable_mechanism",
                state_name="renewable_mechanism_active",
                state_value=1.0,
                source_file=source_file,
                note="新能源机制电价制度自 2026-01-01 起进入执行阶段。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=execution_effective,
                effective_end=None,
                rule_type="renewable_mechanism_execution",
                scope=scope,
                state_group="renewable_mechanism",
                state_name="mechanism_stage_label",
                state_value="执行期",
                source_file=source_file,
                note="新能源机制电价制度执行阶段标签。",
            )

        if "完善2026年度电力中长期交易价格机制" in path.name:
            linkage_effective = pd.Timestamp("2026-02-01 00:00:00")
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=linkage_effective,
                effective_end=None,
                rule_type="lt_price_linkage",
                scope=scope,
                state_group="lt_price_linkage",
                state_name="lt_price_linked_active",
                state_value=1.0,
                source_file=source_file,
                note="2026-02 起年度多月交易进入固定价与联动价组合口径。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=linkage_effective,
                effective_end=None,
                rule_type="lt_price_linkage",
                scope=scope,
                state_group="lt_price_linkage",
                state_name="fixed_price_ratio_max",
                state_value=0.4,
                source_file=source_file,
                note="固定价比例上限 40%。",
            )
            _add_rule(
                rule_rows,
                policy_name=policy_name,
                publish_time=publish_time,
                effective_start=linkage_effective,
                effective_end=None,
                rule_type="lt_price_linkage",
                scope=scope,
                state_group="lt_price_linkage",
                state_name="linked_price_ratio_min",
                state_value=0.6,
                source_file=source_file,
                note="联动价比例下限 60%。",
            )

    inventory = pd.DataFrame(inventory_rows).sort_values(["publish_time", "file_name"], na_position="last").reset_index(drop=True)

    if mechanism_xlsx_summary and mechanism_result_file is not None:
        for state_name, value, note in [
            ("mechanism_price_floor", mechanism_xlsx_summary.get("mechanism_price_floor"), "机制电价下限代理取竞价结果最小值。"),
            ("mechanism_price_ceiling", mechanism_xlsx_summary.get("mechanism_price_ceiling"), "机制电价上限代理取竞价结果最大值。"),
            ("mechanism_volume_ratio_max", mechanism_xlsx_summary.get("mechanism_volume_ratio_max"), "机制电量比例上限取公布结果最大值。"),
            ("mechanism_exec_term_years", mechanism_xlsx_summary.get("mechanism_exec_term_years"), "机制电价执行期限取公布结果最大值。"),
        ]:
            if value is None:
                continue
            _add_rule(
                rule_rows,
                policy_name="湖南省2025年度新能源增量项目机制电价竞价结果公布表",
                publish_time=pd.Timestamp("2025-12-25 00:00:00"),
                effective_start=pd.Timestamp("2026-01-01 00:00:00"),
                effective_end=None,
                rule_type="renewable_mechanism_execution",
                scope="湖南省新能源增量项目",
                state_group="renewable_mechanism",
                state_name=state_name,
                state_value=value,
                source_file=mechanism_result_file,
                note=note,
            )

    failures = pd.DataFrame(failure_rows)
    if failures.empty:
        failures = pd.DataFrame(columns=["source_file", "error"])

    rule_table = pd.DataFrame(rule_rows)
    if rule_table.empty:
        rule_table = pd.DataFrame(
            columns=[
                "rule_id",
                "policy_name",
                "publish_time",
                "effective_start",
                "effective_end",
                "rule_type",
                "scope",
                "state_group",
                "state_name",
                "state_value",
                "source_file",
                "note",
            ]
        )
    else:
        rule_table = rule_table.sort_values(["effective_start", "rule_id"], na_position="last").reset_index(drop=True)

    return PolicyParseResult(inventory=inventory, rule_table=rule_table, failures=failures)
